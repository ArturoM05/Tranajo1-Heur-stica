"""
Combina archivos xlsx parciales (por grupo de instancias) en los archivos finales.
Busca archivos con patrón NWJSSP_ArturoMurgueytio_<Algo>_<grupo>.xlsx
y los une en NWJSSP_ArturoMurgueytio_<Algo>.xlsx
"""
import glob
import os
import openpyxl
from copy import copy


ALGORITHMS = ["Constructivo", "GRASP", "SimulatedAnnealing"]


def merge_workbooks(partial_files, output_file):
    """Combina hojas de varios xlsx parciales en uno solo."""
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for pf in sorted(partial_files):
        wb_in = openpyxl.load_workbook(pf)
        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            ws_out = wb_out.create_sheet(sheet_name)
            for row in ws_in.iter_rows():
                for cell in row:
                    new_cell = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.alignment = copy(cell.alignment)
            for col_letter, dim in ws_in.column_dimensions.items():
                ws_out.column_dimensions[col_letter].width = dim.width
        wb_in.close()

    if wb_out.sheetnames:
        wb_out.save(output_file)
        print(f"  -> {output_file} ({len(wb_out.sheetnames)} hojas)")
    else:
        print(f"  -> {output_file}: sin datos parciales")
    wb_out.close()


def main():
    print("Combinando resultados parciales...")
    for algo in ALGORITHMS:
        pattern = f"NWJSSP_ArturoMurgueytio_{algo}_*.xlsx"
        partial_files = glob.glob(pattern)
        if not partial_files:
            print(f"  {algo}: no se encontraron archivos parciales ({pattern})")
            continue
        print(f"  {algo}: encontrados {len(partial_files)} archivos parciales")
        output = f"NWJSSP_ArturoMurgueytio_{algo}.xlsx"
        merge_workbooks(partial_files, output)
    print("Listo.")


if __name__ == "__main__":
    main()
