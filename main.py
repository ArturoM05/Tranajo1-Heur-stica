"""
Script principal para ejecutar los tres algoritmos de NWJSSP
Genera archivos de resultados en formato Excel según especificaciones del curso

Algoritmos implementados:
1. Constructivo: Greedy determinístico
2. GRASP Simplificado: Construcciones aleatorias con RCL (sin 2-OPT)
3. Simulated Annealing: Enfriamiento progresivo
"""

import os
import glob
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from constructive import ConstructiveAlgorithm
from grasp import GRASP
from simulated_annealing import SimulatedAnnealing
from read_instances import read_nwjssp_instance, calculate_flow_time


# ====== PARÁMETROS DE LOS ALGORITMOS ======

# ALGORITMO CONSTRUCTIVO
CONSTRUCTIVE_PARAMS = {}

# ALGORITMO GRASP SIMPLIFICADO - Más iteraciones para acercarse al óptimo
GRASP_ALPHA = 0.25      # RCL del 25% mejor para equilibrar calidad y diversidad
GRASP_NSOL = 200        # Más construcciones para aumentar probabilidad de buenas soluciones

# ALGORITMO SIMULATED ANNEALING - Búsqueda más profunda
SA_INITIAL_TEMP = 1000.0    # Mayor temperatura inicial para explorar más el espacio
SA_COOLING_RATE = 0.98      # Enfriamiento más lento para refinar la búsqueda
SA_ITERATIONS_PER_TEMP = 100 # Más intentos por nivel de temperatura

# ====== DIRECTORIO DE INSTANCIAS ======
INSTANCES_DIR = "instances"  # Directorio con archivos .txt de instancias


def create_results_workbook(algorithm_name):
    """
    Crea un nuevo workbook para guardar resultados
    
    Args:
        algorithm_name: nombre del algoritmo
        
    Returns:
        workbook: objeto Workbook de openpyxl
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def get_column_letter(col_num):
    """
    Convierte número de columna a letra Excel válida
    Args:
        col_num: número de columna (0-indexado)
    Returns:
        letra de columna (A, B, C, ..., Z, AA, AB, ...)
    """
    col_letter = ""
    col = col_num
    while col >= 0:
        col_letter = chr(65 + (col % 26)) + col_letter
        col = col // 26 - 1
        if col < 0:
            break
    return col_letter


def add_results_sheet(workbook, instance_name, flow_time, computation_time, job_start_times):
    """
    Añade una hoja de resultados para una instancia
    
    Args:
        workbook: objeto Workbook
        instance_name: nombre de la instancia
        flow_time: valor de la función objetivo (makespan)
        computation_time: tiempo de cómputo en milisegundos
        job_start_times: lista con tiempos de inicio de cada trabajo
    """
    # Crear nueva hoja
    ws = workbook.create_sheet(instance_name)
    
    # Primera fila: Z (makespan) y tiempo de cómputo
    ws['A1'] = int(flow_time)
    ws['B1'] = int(round(computation_time))
    
    # Segunda fila: tiempos de inicio de trabajos
    for idx, start_time in enumerate(job_start_times):
        col_letter = get_column_letter(idx)
        cell_ref = f'{col_letter}2'
        ws[cell_ref] = int(start_time)
    
    # Formateo básico
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in [ws['A1'], ws['B1']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for idx in range(len(job_start_times)):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 12


def run_constructive_algorithm(n, m, operations, release_dates):
    """
    Ejecuta el algoritmo Constructivo
    
    Greedy determinístico que ordena trabajos por tiempo total
    
    Args:
        n, m, operations, release_dates: datos del problema
        
    Returns:
        solution, flow_time, computation_time
    """
    algo = ConstructiveAlgorithm(n, m, operations, release_dates)
    solution, flow_time, computation_time = algo.solve()
    return solution, flow_time, computation_time


def run_grasp_algorithm(n, m, operations, release_dates, alpha=GRASP_ALPHA, nsol=GRASP_NSOL):
    """
    Ejecuta el algoritmo GRASP Simplificado
    
    Construcciones aleatorias usando RCL (Lista Restringida de Candidatos)
    Sin búsqueda local 2-OPT para mejor velocidad
    
    Parámetros:
    - alpha=0.25: usa una RCL del 25% mejor para combinar greedy y diversidad
    - nsol=200: genera más construcciones aleatorias para quedarse con la mejor
    
    Args:
        n, m, operations, release_dates: datos del problema
        alpha: parámetro de restricción de candidatos
        nsol: número de soluciones a generar
        
    Returns:
        solution, flow_time, computation_time
    """
    algo = GRASP(n, m, operations, release_dates, alpha=alpha, nsol=nsol)
    solution, flow_time, computation_time = algo.solve()
    return solution, flow_time, computation_time


def run_simulated_annealing_algorithm(n, m, operations, release_dates, 
                                      initial_temp=SA_INITIAL_TEMP,
                                      cooling_rate=SA_COOLING_RATE,
                                      iterations_per_temp=SA_ITERATIONS_PER_TEMP):
    """
    Ejecuta el algoritmo Simulated Annealing
    
    Enfriamiento progresivo: empieza aceptando soluciones malas (explora)
    y progresivamente solo acepta mejoras (explota)
    
    Parámetros:
    - initial_temp=1000: Temperatura inicial más alta para explorar más
    - cooling_rate=0.98: Enfría 2% por iteración para una búsqueda más gradual
    - iterations_per_temp=100: Más intentos en cada nivel de temperatura
    
    Args:
        n, m, operations, release_dates: datos del problema
        initial_temp: temperatura inicial
        cooling_rate: tasa de enfriamiento
        iterations_per_temp: iteraciones por temperatura
        
    Returns:
        solution, flow_time, computation_time
    """
    algo = SimulatedAnnealing(n, m, operations, release_dates,
                              initial_temp=initial_temp,
                              cooling_rate=cooling_rate,
                              iterations_per_temp=iterations_per_temp)
    solution, flow_time, computation_time = algo.solve()
    return solution, flow_time, computation_time


def process_instance(instance_file, algorithm_func, **algo_kwargs):
    """
    Procesa una instancia con un algoritmo específico
    
    Args:
        instance_file: ruta del archivo de instancia
        algorithm_func: función del algoritmo a ejecutar
        **algo_kwargs: parámetros adicionales del algoritmo
        
    Returns:
        tuple: (nombre_instancia, solution, flow_time, computation_time) o None si hay error
    """
    try:
        # Leer instancia
        n, m, operations, release_dates, L = read_nwjssp_instance(instance_file)
        
        # Ejecutar algoritmo
        solution, flow_time, computation_time = algorithm_func(n, m, operations, release_dates, **algo_kwargs)
        
        # Obtener nombre de instancia
        instance_name = os.path.splitext(os.path.basename(instance_file))[0]
        
        return instance_name, solution, flow_time, computation_time
    
    except Exception as e:
        print(f"Error procesando {instance_file}: {str(e)}")
        return None


def main():
    """
    Función principal: ejecuta todos los algoritmos en todas las instancias
    
    Genera tres archivos Excel:
    1. NWJSSP_ArturoMurgueytio_Constructivo.xlsx
    2. NWJSSP_ArturoMurgueytio_GRASP.xlsx
    3. NWJSSP_ArturoMurgueytio_SimulatedAnnealing.xlsx
    """
    print("=" * 70)
    print("NWJSSP - Solver: Constructivo, GRASP Simplificado, Simulated Annealing")
    print("=" * 70)
    print()
    
    # Buscar archivos de instancias
    instance_files = glob.glob(os.path.join(INSTANCES_DIR, "*.txt"))
    instance_files.sort()
    
    if not instance_files:
        print(f"Error: No se encontraron archivos de instancias en {INSTANCES_DIR}")
        return
    
    print(f"Se encontraron {len(instance_files)} instancias")
    print()
    
    # ====== ALGORITMO 1: CONSTRUCTIVO ======
    print("1. Ejecutando Algoritmo CONSTRUCTIVO (Greedy determinístico)...")
    print("-" * 70)
    wb_constructive = create_results_workbook("Constructive")
    
    constructive_time_total = 0
    constructive_count = 0
    
    for instance_file in instance_files:
        result = process_instance(instance_file, run_constructive_algorithm)
        if result:
            instance_name, solution, flow_time, computation_time = result
            print(f"  {instance_name:30s} | Z={flow_time:12.0f} | Tiempo={computation_time:8.2f}ms")
            constructive_time_total += computation_time
            constructive_count += 1
            add_results_sheet(wb_constructive, instance_name, flow_time, computation_time, solution)
    
    output_file_constructive = "NWJSSP_ArturoMurgueytio_Constructivo.xlsx"
    wb_constructive.save(output_file_constructive)
    print(f"\n✓ Resultados guardados en {output_file_constructive}")
    print(f"  Tiempo total: {constructive_time_total/1000:.2f}s, Promedio: {constructive_time_total/constructive_count:.2f}ms")
    print()
    
    # ====== ALGORITMO 2: GRASP SIMPLIFICADO ======
    print("2. Ejecutando Algoritmo GRASP SIMPLIFICADO (Construcciones aleatorias)...")
    print("-" * 70)
    wb_grasp = create_results_workbook("GRASP")
    
    grasp_time_total = 0
    grasp_count = 0
    
    for instance_file in instance_files:
        result = process_instance(instance_file, run_grasp_algorithm, 
                                 alpha=GRASP_ALPHA, nsol=GRASP_NSOL)
        if result:
            instance_name, solution, flow_time, computation_time = result
            print(f"  {instance_name:30s} | Z={flow_time:12.0f} | Tiempo={computation_time:8.2f}ms")
            grasp_time_total += computation_time
            grasp_count += 1
            add_results_sheet(wb_grasp, instance_name, flow_time, computation_time, solution)
    
    output_file_grasp = "NWJSSP_ArturoMurgueytio_GRASP.xlsx"
    wb_grasp.save(output_file_grasp)
    print(f"\n✓ Resultados guardados en {output_file_grasp}")
    print(f"  Tiempo total: {grasp_time_total/1000:.2f}s, Promedio: {grasp_time_total/grasp_count:.2f}ms")
    print()
    
    # ====== ALGORITMO 3: SIMULATED ANNEALING ======
    print("3. Ejecutando Algoritmo SIMULATED ANNEALING (Enfriamiento progresivo)...")
    print("-" * 70)
    wb_sa = create_results_workbook("SimulatedAnnealing")
    
    sa_time_total = 0
    sa_count = 0
    
    for instance_file in instance_files:
        result = process_instance(instance_file, run_simulated_annealing_algorithm,
                                 initial_temp=SA_INITIAL_TEMP,
                                 cooling_rate=SA_COOLING_RATE,
                                 iterations_per_temp=SA_ITERATIONS_PER_TEMP)
        if result:
            instance_name, solution, flow_time, computation_time = result
            print(f"  {instance_name:30s} | Z={flow_time:12.0f} | Tiempo={computation_time:8.2f}ms")
            sa_time_total += computation_time
            sa_count += 1
            add_results_sheet(wb_sa, instance_name, flow_time, computation_time, solution)
    
    output_file_sa = "NWJSSP_ArturoMurgueytio_SimulatedAnnealing.xlsx"
    wb_sa.save(output_file_sa)
    print(f"\n✓ Resultados guardados en {output_file_sa}")
    print(f"  Tiempo total: {sa_time_total/1000:.2f}s, Promedio: {sa_time_total/sa_count:.2f}ms")
    print()
    
    # ====== RESUMEN COMPARATIVO ======
    print("=" * 70)
    print("RESUMEN COMPARATIVO")
    print("=" * 70)
    print(f"{'Algoritmo':<30} | {'Instancias':<12} | {'Tiempo Total':<15} | {'Promedio':<12}")
    print("-" * 70)
    print(f"{'Constructivo':<30} | {constructive_count:<12} | {constructive_time_total/1000:>6.2f}s     | {constructive_time_total/constructive_count:>6.2f}ms")
    print(f"{'GRASP Simplificado':<30} | {grasp_count:<12} | {grasp_time_total/1000:>6.2f}s     | {grasp_time_total/grasp_count:>6.2f}ms")
    print(f"{'Simulated Annealing':<30} | {sa_count:<12} | {sa_time_total/1000:>6.2f}s     | {sa_time_total/sa_count:>6.2f}ms")
    print("=" * 70)
    print()
    print("✓ Ejecución completada exitosamente")
    print()
    print("Archivos generados:")
    print(f"  1. {output_file_constructive}")
    print(f"  2. {output_file_grasp}")
    print(f"  3. {output_file_sa}")
    print()


if __name__ == "__main__":
    main()